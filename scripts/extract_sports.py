"""
Extract sport-level handle data from Indiana's raw Excel revenue reports.

Reads Sheet 7 ("7 SW Tax Summary" or "Sheet7") from each IN_*-Revenue.xlsx file,
finds the "Statewide Handle by Sport" section, and extracts monthly handle values
for each sport (Football, Basketball, Baseball, Parlay, Other).

Output: data/processed/IN_sports.csv with columns: period_end, sport, handle
"""

import os
import re
import calendar
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd


RAW_DIR = Path(__file__).resolve().parent.parent / "data" / "raw" / "IN"
OUTPUT_PATH = Path(__file__).resolve().parent.parent / "data" / "processed" / "IN_sports.csv"

SPORT_NAMES = {"Football", "Basketball", "Baseball", "Parlay", "Other"}


def find_sheet7(workbook):
    """Find the sheet containing the SW Tax Summary (Sheet 7)."""
    # Try exact matches first
    for name in ["7 SW Tax Summary", "Sheet7"]:
        if name in workbook.sheetnames:
            return workbook[name]

    # Try partial match with '7'
    for name in workbook.sheetnames:
        if "7" in name and ("SW" in name or "Tax" in name or "Sport" in name):
            return workbook[name]

    # Fall back to the 7th sheet (index 6) if it exists
    if len(workbook.sheetnames) >= 7:
        return workbook[workbook.sheetnames[6]]

    return None


def find_month_column_and_sport_rows(ws):
    """
    Find the column containing monthly handle data and the rows with sport data.

    Strategy:
    1. Locate the cell containing "Month" text in the sport-section header row.
    2. Sport data rows start immediately after, with sport names in column A.
    3. Stop when we hit "TOTAL" or run out of known sport names.

    Returns:
        month_col (int): 1-based column index for monthly handle values
        sport_rows (list of tuples): [(row_number, sport_name), ...]
    """
    month_cell = None

    # Scan entire sheet for the "Month" header cell
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value and str(cell.value).strip() == "Month":
                month_cell = cell
                break
        if month_cell is not None:
            break

    if month_cell is None:
        return None, []

    month_col = month_cell.column  # 1-based column index
    header_row = month_cell.row

    # Sport rows follow the header row. Scan downward from header_row + 1.
    sport_rows = []
    for row_num in range(header_row + 1, header_row + 20):
        cell_a = ws.cell(row=row_num, column=1)
        if cell_a.value is None:
            continue
        name = str(cell_a.value).strip()
        if name.upper() == "TOTAL":
            break
        if name in SPORT_NAMES:
            sport_rows.append((row_num, name))

    return month_col, sport_rows


def parse_period_end(filename):
    """
    Extract year-month from filename and return the last day of that month.
    Filename format: IN_YYYY-MM-Revenue.xlsx
    """
    match = re.search(r"(\d{4})-(\d{2})", filename)
    if not match:
        return None
    year, month = int(match.group(1)), int(match.group(2))
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, last_day)


def extract_sports_from_file(filepath):
    """
    Extract sport-level handle data from a single Excel file.

    Returns:
        list of dicts with keys: period_end, sport, handle
    """
    filename = os.path.basename(filepath)
    period_end = parse_period_end(filename)
    if period_end is None:
        print(f"  WARNING: Could not parse date from {filename}, skipping.")
        return []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  WARNING: Could not open {filename}: {e}")
        return []

    ws = find_sheet7(wb)
    if ws is None:
        print(f"  WARNING: No Sheet 7 found in {filename}, skipping.")
        wb.close()
        return []

    month_col, sport_rows = find_month_column_and_sport_rows(ws)
    if month_col is None or not sport_rows:
        print(f"  WARNING: Could not find sport section in {filename}, skipping.")
        wb.close()
        return []

    records = []
    for row_num, sport_name in sport_rows:
        handle_cell = ws.cell(row=row_num, column=month_col)
        handle_value = handle_cell.value

        # Handle missing / non-numeric values
        if handle_value is None:
            handle_value = 0.0
        else:
            try:
                handle_value = float(handle_value)
            except (ValueError, TypeError):
                print(
                    f"  WARNING: Non-numeric handle for {sport_name} in {filename}: {handle_value}"
                )
                handle_value = 0.0

        records.append(
            {
                "period_end": period_end.isoformat(),
                "sport": sport_name,
                "handle": handle_value,
            }
        )

    wb.close()
    return records


def main():
    # Collect all IN_*-Revenue.xlsx files
    files = sorted(RAW_DIR.glob("IN_*-Revenue.xlsx"))
    if not files:
        print(f"No IN_*-Revenue.xlsx files found in {RAW_DIR}")
        return

    print(f"Found {len(files)} files to process.")

    all_records = []
    for filepath in files:
        records = extract_sports_from_file(filepath)
        if records:
            print(f"  {filepath.name}: extracted {len(records)} sport rows")
        all_records.extend(records)

    if not all_records:
        print("No data extracted.")
        return

    # Build DataFrame and sort: period_end descending, sport alphabetically
    df = pd.DataFrame(all_records)
    df = df.sort_values(
        by=["period_end", "sport"], ascending=[False, True]
    ).reset_index(drop=True)

    # Write CSV
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUTPUT_PATH, index=False)
    print(f"\nWrote {len(df)} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
