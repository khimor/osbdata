#!/usr/bin/env python3
"""
Compile all tier research results into master output files.
Run this after all tier agents have completed.
"""
import json
import os
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESEARCH_DIR = Path("/Users/nosherzapoo/Desktop/claude/osb-trackerv0/research")

def load_all_results():
    """Load and merge all tier results."""
    all_states = {}
    for tier_file in ["tier1_results.json", "tier2_results.json", "tier3_results.json",
                      "tier4_results.json", "tier5_results.json"]:
        path = RESEARCH_DIR / tier_file
        if path.exists():
            with open(path) as f:
                data = json.load(f)
                all_states.update(data)
            print(f"  Loaded {tier_file}: {len(data)} states")
        else:
            print(f"  WARNING: {tier_file} not found!")
    return all_states

def load_tax_research():
    """Load supplementary tax/GGR research."""
    path = RESEARCH_DIR / "tax_and_ggr_research.json"
    if path.exists():
        with open(path) as f:
            return json.load(f)
    return {}

def merge_tax_data(all_states, tax_data):
    """Merge tax research into state data, filling gaps."""
    for state_code, tax_info in tax_data.items():
        if state_code in all_states:
            state = all_states[state_code]
            # Fill in missing fields from tax research
            if not state.get("online_tax_rate") or state["online_tax_rate"] in ["UNKNOWN", "", None]:
                state["online_tax_rate"] = tax_info.get("online_tax_rate", "UNKNOWN")
            if not state.get("retail_tax_rate") or state["retail_tax_rate"] in ["UNKNOWN", "", None]:
                state["retail_tax_rate"] = tax_info.get("retail_tax_rate", "UNKNOWN")
            if not state.get("tax_basis") or state["tax_basis"] in ["UNKNOWN", "", None]:
                state["tax_basis"] = tax_info.get("tax_basis", "UNKNOWN")
            if not state.get("ggr_definition") or state["ggr_definition"] in ["UNKNOWN", "", None]:
                state["ggr_definition"] = tax_info.get("ggr_term_used", "UNKNOWN")
            if not state.get("ggr_formula") or state["ggr_formula"] in ["UNKNOWN", "", None]:
                state["ggr_formula"] = tax_info.get("ggr_formula", "UNKNOWN")
            if not state.get("promo_deduction") or state["promo_deduction"] in ["UNKNOWN", "", None]:
                state["promo_deduction"] = tax_info.get("promo_deduction_details", "UNKNOWN")
            if not state.get("launch_date") or state["launch_date"] in ["UNKNOWN", "", None]:
                state["launch_date"] = tax_info.get("launch_date", "UNKNOWN")
            if tax_info.get("special_notes"):
                existing_notes = state.get("notes", "") or ""
                state["notes"] = f"{existing_notes} | Tax notes: {tax_info['special_notes']}"
        else:
            # State not in tier results, create minimal entry from tax data
            all_states[state_code] = {
                "state_name": tax_info.get("state_name", state_code),
                "regulatory_body": "UNKNOWN",
                "url": "UNKNOWN",
                "url_status": "not_checked",
                "launch_date": tax_info.get("launch_date", "UNKNOWN"),
                "reporting_frequency": "UNKNOWN",
                "file_formats": [],
                "has_archive": False,
                "archive_depth": "UNKNOWN",
                "requires_js": False,
                "operator_breakdown": False,
                "operators": [],
                "sport_breakdown": False,
                "sport_categories": [],
                "retail_online_split": False,
                "data_fields": [],
                "ggr_definition": tax_info.get("ggr_term_used", "UNKNOWN"),
                "ggr_formula": tax_info.get("ggr_formula", "UNKNOWN"),
                "promo_deduction": tax_info.get("promo_deduction_details", "UNKNOWN"),
                "online_tax_rate": tax_info.get("online_tax_rate", "UNKNOWN"),
                "retail_tax_rate": tax_info.get("retail_tax_rate", "UNKNOWN"),
                "tax_basis": tax_info.get("tax_basis", "UNKNOWN"),
                "scraping_difficulty": "UNKNOWN",
                "scraping_approach": "UNKNOWN",
                "scraping_time_estimate": "UNKNOWN",
                "sample_files_downloaded": [],
                "notes": tax_info.get("special_notes", ""),
                "log": "Data from supplementary tax research only"
            }
    return all_states

def safe_get(d, key, default="UNKNOWN"):
    """Safely get a value from dict, returning default for None/empty."""
    val = d.get(key, default)
    if val is None or val == "" or val == []:
        return default
    return val

def create_master_spreadsheet(all_states):
    """Create the master Excel spreadsheet with all 7 sheets."""
    wb = Workbook()

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header_row(ws, num_cols, row=1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def auto_width(ws, min_width=12, max_width=40):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

    # ===== Sheet 1: State Overview =====
    ws1 = wb.active
    ws1.title = "State Overview"
    headers1 = [
        "State Code", "State Name", "Regulatory Body", "URL", "URL Status",
        "Launch Date", "Frequency", "Format", "Operator Breakdown",
        "Sport Breakdown", "Retail/Online Split", "Online Tax Rate",
        "Retail Tax Rate", "Tax Basis", "Scraping Difficulty",
        "Est. Scraper Build Time", "Archive Depth", "Sample File Saved",
        "Tier", "Notes"
    ]
    ws1.append(headers1)
    style_header_row(ws1, len(headers1))

    # Sort states by code
    sorted_states = sorted(all_states.items())

    tier_map = {
        "NY": 1, "IL": 1, "PA": 1, "NJ": 1, "OH": 1, "MI": 1,
        "AZ": 2, "IN": 2, "MA": 2, "MD": 2, "VA": 2, "CO": 2,
        "CT": 3, "KY": 3, "TN": 3, "LA": 3, "NC": 3, "KS": 3, "IA": 3,
        "WV": 4, "ME": 4, "NH": 4, "RI": 4, "WY": 4, "MO": 4,
        "MT": 5, "OR": 5, "SD": 5, "DE": 5, "DC": 5, "AR": 5, "VT": 5, "MS": 5, "NV": 5
    }

    for code, data in sorted_states:
        formats = safe_get(data, "file_formats", [])
        if isinstance(formats, list):
            formats = ", ".join(formats)

        sample_files = safe_get(data, "sample_files_downloaded", [])
        has_samples = "Y" if sample_files and len(sample_files) > 0 else "N"

        row = [
            code,
            safe_get(data, "state_name"),
            safe_get(data, "regulatory_body"),
            safe_get(data, "url"),
            safe_get(data, "url_status"),
            safe_get(data, "launch_date"),
            safe_get(data, "reporting_frequency"),
            formats,
            "Y" if data.get("operator_breakdown") else "N",
            "Y" if data.get("sport_breakdown") else "N",
            "Y" if data.get("retail_online_split") else "N",
            safe_get(data, "online_tax_rate"),
            safe_get(data, "retail_tax_rate"),
            safe_get(data, "tax_basis"),
            safe_get(data, "scraping_difficulty"),
            safe_get(data, "scraping_time_estimate"),
            safe_get(data, "archive_depth"),
            has_samples,
            tier_map.get(code, "?"),
            safe_get(data, "notes", "")[:200]
        ]
        ws1.append(row)

    auto_width(ws1)

    # ===== Sheet 2: Data Fields by State =====
    ws2 = wb.create_sheet("Data Fields by State")

    # Collect all unique field names across states
    all_field_names = set()
    standard_fields = [
        "Handle (Total Wagers)", "Online Handle", "Retail Handle",
        "Gross Revenue (before promos)", "Promo Credits / Free Bets",
        "Net Revenue (after promos) / AGR", "Payouts / Winnings Paid",
        "Hold %", "Tax Paid", "Tax Rate", "Operator Name",
        "Sport Category", "Bet Type", "New Accounts / Registrations",
        "Number of Bets / Transactions", "Federal Excise Tax"
    ]
    all_field_names.update(standard_fields)

    for code, data in sorted_states:
        fields = safe_get(data, "data_fields", [])
        if isinstance(fields, list):
            for f in fields:
                if isinstance(f, dict):
                    all_field_names.add(f.get("name", ""))
                elif isinstance(f, str):
                    all_field_names.add(f)

    all_field_names.discard("")
    sorted_fields = sorted(all_field_names)

    headers2 = ["State Code"] + sorted_fields
    ws2.append(headers2)
    style_header_row(ws2, len(headers2))

    for code, data in sorted_states:
        state_fields = set()
        fields = safe_get(data, "data_fields", [])
        if isinstance(fields, list):
            for f in fields:
                if isinstance(f, dict):
                    state_fields.add(f.get("name", "").lower())
                elif isinstance(f, str):
                    state_fields.add(f.lower())

        row = [code]
        for field in sorted_fields:
            # Check if this field or something similar exists
            field_lower = field.lower()
            found = any(field_lower in sf or sf in field_lower for sf in state_fields)
            if found:
                row.append("Y")
            elif data.get("operator_breakdown") and "operator" in field_lower:
                row.append("Y")
            elif data.get("sport_breakdown") and "sport" in field_lower:
                row.append("Y")
            else:
                row.append("N")
        ws2.append(row)

    auto_width(ws2, min_width=8, max_width=25)

    # ===== Sheet 3: Operator Name Mapping =====
    ws3 = wb.create_sheet("Operator Name Mapping")
    headers3 = ["State", "Name As Reported", "Standardized Name"]
    ws3.append(headers3)
    style_header_row(ws3, len(headers3))

    # Build operator mapping
    operator_standardization = {
        "fanduel": "FanDuel",
        "fanduel sportsbook": "FanDuel",
        "fanduel group": "FanDuel",
        "flutter/fanduel": "FanDuel",
        "draftkings": "DraftKings",
        "draftkings sportsbook": "DraftKings",
        "crown gaming / draftkings": "DraftKings",
        "crown gaming/draftkings": "DraftKings",
        "betmgm": "BetMGM",
        "bet mgm": "BetMGM",
        "roar digital": "BetMGM",
        "borgata": "BetMGM",
        "caesars": "Caesars Sportsbook",
        "caesars sportsbook": "Caesars Sportsbook",
        "william hill": "Caesars Sportsbook",
        "caesars interactive": "Caesars Sportsbook",
        "espn bet": "ESPN Bet",
        "penn entertainment / espn bet": "ESPN Bet",
        "penn entertainment": "ESPN Bet",
        "wynn interactive": "ESPN Bet",
        "wynnbet": "ESPN Bet",
        "barstool sportsbook": "ESPN Bet",
        "barstool": "ESPN Bet",
        "betrivers": "BetRivers",
        "rush street interactive": "BetRivers",
        "rivers casino": "BetRivers",
        "sugarhouse": "BetRivers",
        "pointsbet": "Fanatics Sportsbook",
        "fanatics": "Fanatics Sportsbook",
        "fanatics sportsbook": "Fanatics Sportsbook",
        "bet365": "bet365",
        "hard rock": "Hard Rock Bet",
        "hard rock bet": "Hard Rock Bet",
        "seminole hard rock": "Hard Rock Bet",
        "betparx": "Parx",
        "parx casino": "Parx",
        "parx": "Parx",
        "resorts digital": "Resorts Digital",
        "fox bet": "Discontinued",
        "foxbet": "Discontinued",
        "ballybet": "Bally Bet",
        "bally bet": "Bally Bet",
        "bally's": "Bally Bet",
    }

    for code, data in sorted_states:
        operators = safe_get(data, "operators", [])
        if isinstance(operators, list):
            for op in operators:
                if isinstance(op, str) and op.strip():
                    std = operator_standardization.get(op.strip().lower(), op.strip())
                    ws3.append([code, op.strip(), std])

    auto_width(ws3)

    # ===== Sheet 4: GGR Definitions =====
    ws4 = wb.create_sheet("GGR Definitions")
    headers4 = [
        "State", "GGR Term Used", "Formula Description",
        "Promo Deduction Allowed", "Promo Deduction Limit",
        "Taxable Base Definition", "Notes"
    ]
    ws4.append(headers4)
    style_header_row(ws4, len(headers4))

    for code, data in sorted_states:
        promo = safe_get(data, "promo_deduction", "UNKNOWN")
        promo_allowed = "UNKNOWN"
        if isinstance(promo, bool):
            promo_allowed = "Y" if promo else "N"
        elif isinstance(promo, str):
            promo_lower = promo.lower()
            if "not" in promo_lower or "no" == promo_lower or "false" in promo_lower:
                promo_allowed = "N"
            elif "yes" in promo_lower or "allowed" in promo_lower or "true" in promo_lower:
                promo_allowed = "Y"
            elif "partial" in promo_lower or "sliding" in promo_lower or "limited" in promo_lower:
                promo_allowed = "Partial"
            else:
                promo_allowed = promo[:50]

        row = [
            code,
            safe_get(data, "ggr_definition"),
            safe_get(data, "ggr_formula"),
            promo_allowed,
            safe_get(data, "promo_deduction", "")[:100],
            safe_get(data, "tax_basis"),
            safe_get(data, "notes", "")[:200]
        ]
        ws4.append(row)

    auto_width(ws4)

    # ===== Sheet 5: Sport Categories =====
    ws5 = wb.create_sheet("Sport Categories")
    headers5 = ["State", "Sport Category", "Is Handle Reported", "Is Revenue Reported"]
    ws5.append(headers5)
    style_header_row(ws5, len(headers5))

    for code, data in sorted_states:
        if data.get("sport_breakdown"):
            categories = safe_get(data, "sport_categories", [])
            if isinstance(categories, list):
                for cat in categories:
                    if isinstance(cat, str):
                        ws5.append([code, cat, "Y", "Y"])
                    elif isinstance(cat, dict):
                        ws5.append([
                            code,
                            cat.get("name", cat.get("category", str(cat))),
                            "Y" if cat.get("handle_reported", True) else "N",
                            "Y" if cat.get("revenue_reported", True) else "N"
                        ])
            if not categories or len(categories) == 0:
                ws5.append([code, "Has sport breakdown but categories not documented", "UNKNOWN", "UNKNOWN"])

    auto_width(ws5)

    # ===== Sheet 6: Scraping Plan =====
    ws6 = wb.create_sheet("Scraping Plan")
    headers6 = [
        "State", "Recommended Tool", "URL Pattern for Downloads",
        "Auth/Headers Needed", "Anti-Scraping Notes",
        "Est. Build Time", "Priority Order"
    ]
    ws6.append(headers6)
    style_header_row(ws6, len(headers6))

    priority_order = 1
    for tier_states in [
        ["NY", "IL", "PA", "NJ", "OH", "MI"],
        ["AZ", "IN", "MA", "MD", "VA", "CO"],
        ["CT", "KY", "TN", "LA", "NC", "KS", "IA"],
        ["WV", "ME", "NH", "RI", "WY", "MO"],
        ["MT", "OR", "SD", "DE", "DC", "AR", "VT", "MS", "NV"]
    ]:
        for code in tier_states:
            data = all_states.get(code, {})
            row = [
                code,
                safe_get(data, "scraping_approach"),
                safe_get(data, "url"),
                "Standard headers" if not data.get("requires_js") else "Playwright needed",
                safe_get(data, "notes", "")[:100],
                safe_get(data, "scraping_time_estimate"),
                priority_order
            ]
            ws6.append(row)
            priority_order += 1

    auto_width(ws6)

    # ===== Sheet 7: Sample Data Preview =====
    ws7 = wb.create_sheet("Sample Data Preview")
    headers7 = ["State", "Source File", "Data Preview (first rows)"]
    ws7.append(headers7)
    style_header_row(ws7, len(headers7))

    for code, data in sorted_states:
        sample_files = safe_get(data, "sample_files_downloaded", [])
        if isinstance(sample_files, list) and len(sample_files) > 0:
            for sf in sample_files[:2]:  # Max 2 files per state
                if isinstance(sf, str):
                    ws7.append([code, sf, "See downloaded file"])
                elif isinstance(sf, dict):
                    ws7.append([code, sf.get("filename", "unknown"), sf.get("preview", "See file")])
        else:
            ws7.append([code, "No sample files downloaded", ""])

    auto_width(ws7)

    # Save
    output_path = RESEARCH_DIR / "state_data_research.xlsx"
    wb.save(output_path)
    print(f"Saved master spreadsheet to {output_path}")
    return output_path

def create_operator_mapping_json(all_states):
    """Create the operator name mapping JSON."""

    # Standard operator info
    standard_operators = {
        "FanDuel": {
            "parent_company": "Flutter Entertainment",
            "aliases": set(),
            "active_states": set()
        },
        "DraftKings": {
            "parent_company": "DraftKings Inc.",
            "aliases": set(),
            "active_states": set()
        },
        "BetMGM": {
            "parent_company": "Entain/MGM Resorts",
            "aliases": set(),
            "active_states": set()
        },
        "Caesars Sportsbook": {
            "parent_company": "Caesars Entertainment",
            "aliases": set(),
            "active_states": set()
        },
        "ESPN Bet": {
            "parent_company": "PENN Entertainment",
            "aliases": set(),
            "active_states": set(),
            "notes": "Formerly WynnBET, formerly Barstool Sportsbook. PENN rebranded all skins to ESPN Bet."
        },
        "BetRivers": {
            "parent_company": "Rush Street Interactive",
            "aliases": set(),
            "active_states": set()
        },
        "Fanatics Sportsbook": {
            "parent_company": "Fanatics Betting & Gaming",
            "aliases": set(),
            "active_states": set(),
            "notes": "Acquired PointsBet US operations in 2023."
        },
        "bet365": {
            "parent_company": "bet365 Group",
            "aliases": set(),
            "active_states": set()
        },
        "Hard Rock Bet": {
            "parent_company": "Seminole Tribe of Florida / Hard Rock Digital",
            "aliases": set(),
            "active_states": set()
        },
        "Bally Bet": {
            "parent_company": "Bally's Corporation",
            "aliases": set(),
            "active_states": set()
        }
    }

    operator_normalize = {
        "fanduel": "FanDuel",
        "fanduel sportsbook": "FanDuel",
        "fanduel group": "FanDuel",
        "flutter/fanduel": "FanDuel",
        "draftkings": "DraftKings",
        "draftkings sportsbook": "DraftKings",
        "crown gaming / draftkings": "DraftKings",
        "crown gaming/draftkings": "DraftKings",
        "dk": "DraftKings",
        "betmgm": "BetMGM",
        "bet mgm": "BetMGM",
        "roar digital": "BetMGM",
        "borgata": "BetMGM",
        "caesars": "Caesars Sportsbook",
        "caesars sportsbook": "Caesars Sportsbook",
        "william hill": "Caesars Sportsbook",
        "caesars interactive": "Caesars Sportsbook",
        "espn bet": "ESPN Bet",
        "penn entertainment / espn bet": "ESPN Bet",
        "penn entertainment": "ESPN Bet",
        "wynn interactive": "ESPN Bet",
        "wynnbet": "ESPN Bet",
        "barstool sportsbook": "ESPN Bet",
        "barstool": "ESPN Bet",
        "betrivers": "BetRivers",
        "rush street interactive": "BetRivers",
        "rivers": "BetRivers",
        "sugarhouse": "BetRivers",
        "pointsbet": "Fanatics Sportsbook",
        "fanatics": "Fanatics Sportsbook",
        "fanatics sportsbook": "Fanatics Sportsbook",
        "bet365": "bet365",
        "hard rock": "Hard Rock Bet",
        "hard rock bet": "Hard Rock Bet",
        "hard rock digital": "Hard Rock Bet",
        "seminole hard rock": "Hard Rock Bet",
        "ballybet": "Bally Bet",
        "bally bet": "Bally Bet",
        "bally's": "Bally Bet",
    }

    state_specific_mappings = {}

    for code, data in all_states.items():
        operators = data.get("operators", [])
        if isinstance(operators, list):
            for op in operators:
                if isinstance(op, str) and op.strip():
                    op_clean = op.strip()
                    std_name = operator_normalize.get(op_clean.lower(), None)

                    if std_name and std_name in standard_operators:
                        standard_operators[std_name]["aliases"].add(op_clean)
                        standard_operators[std_name]["active_states"].add(code)

                        if op_clean.lower() != std_name.lower():
                            if code not in state_specific_mappings:
                                state_specific_mappings[code] = {}
                            state_specific_mappings[code][op_clean] = std_name

    # Convert sets to sorted lists for JSON serialization
    output = {
        "standardized_names": {},
        "state_specific_mappings": state_specific_mappings
    }

    for name, info in standard_operators.items():
        output["standardized_names"][name] = {
            "aliases": sorted(info["aliases"]),
            "parent_company": info["parent_company"],
            "active_states": sorted(info["active_states"]),
            "notes": info.get("notes", "")
        }

    output_path = RESEARCH_DIR / "operator_name_mapping.json"
    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)
    print(f"Saved operator mapping to {output_path}")
    return output_path

def create_unified_schema_json():
    """Create the unified database schema recommendation."""
    schema = {
        "description": "Recommended database schema that accommodates all US state sports betting revenue data",
        "tables": {
            "states": {
                "columns": {
                    "state_code": "VARCHAR(2) PRIMARY KEY",
                    "state_name": "VARCHAR(50)",
                    "regulatory_body": "VARCHAR(100)",
                    "source_url": "TEXT",
                    "launch_date": "DATE",
                    "reporting_frequency": "VARCHAR(20)",
                    "online_tax_rate": "DECIMAL(5,4)",
                    "retail_tax_rate": "DECIMAL(5,4)",
                    "tax_basis": "VARCHAR(20) -- 'ggr', 'handle', 'agr'",
                    "promo_deduction_allowed": "BOOLEAN",
                    "promo_deduction_cap": "TEXT",
                    "ggr_definition": "TEXT",
                    "data_format": "VARCHAR(50)",
                    "has_operator_breakdown": "BOOLEAN",
                    "has_sport_breakdown": "BOOLEAN",
                    "has_channel_split": "BOOLEAN"
                }
            },
            "monthly_data": {
                "description": "Main data table — one row per state/period/operator combination",
                "columns": {
                    "state_code": "VARCHAR(2)",
                    "period_start": "DATE",
                    "period_end": "DATE",
                    "period_type": "VARCHAR(10) -- 'weekly', 'monthly'",
                    "operator_raw": "VARCHAR(100) -- name as reported by state",
                    "operator_standard": "VARCHAR(100) -- normalized name from mapping",
                    "channel": "VARCHAR(20) -- 'online', 'retail', 'kiosk', 'combined'",
                    "sport_category": "VARCHAR(50) -- NULL if state doesn't break down by sport",
                    "handle": "BIGINT -- total wagers in cents",
                    "gross_revenue": "BIGINT -- handle minus payouts, before promo deductions, in cents",
                    "promo_credits": "BIGINT -- free bets / promotional credits in cents, NULL if not reported",
                    "net_revenue": "BIGINT -- after promo deductions (taxable base in most states), in cents",
                    "payouts": "BIGINT -- winnings paid to bettors, NULL if not reported, in cents",
                    "hold_pct": "DECIMAL(6,4) -- gross_revenue / handle, NULL if not calculable",
                    "tax_paid": "BIGINT -- actual tax remitted, in cents",
                    "num_bets": "INTEGER -- number of wagers, NULL if not reported",
                    "federal_excise_tax": "BIGINT -- NULL if not reported separately, in cents"
                },
                "primary_key": ["state_code", "period_end", "operator_standard", "channel", "sport_category"],
                "notes": "Store currency as cents (BIGINT) to avoid floating point issues. Convert display to dollars in the application layer."
            },
            "scrape_log": {
                "columns": {
                    "id": "SERIAL PRIMARY KEY",
                    "state_code": "VARCHAR(2)",
                    "scrape_timestamp": "TIMESTAMP",
                    "status": "VARCHAR(20) -- 'success', 'partial', 'failed'",
                    "rows_inserted": "INTEGER",
                    "date_range_start": "DATE",
                    "date_range_end": "DATE",
                    "error_message": "TEXT",
                    "source_file": "TEXT"
                }
            }
        },
        "derived_views": {
            "hold_pct": "gross_revenue / NULLIF(handle, 0)",
            "yoy_handle_change": "handle - LAG(handle) OVER (PARTITION BY state_code, operator_standard ORDER BY period_end) for same month prior year",
            "market_share": "handle / SUM(handle) OVER (PARTITION BY state_code, period_end)",
            "effective_tax_rate": "tax_paid / NULLIF(net_revenue, 0)"
        },
        "columns_that_vary_by_state": [
            "promo_credits -- some states don't report this at all",
            "payouts -- only ~60% of states report raw payouts",
            "sport_category -- only IL, MA, and maybe 1-2 others",
            "num_bets -- very few states",
            "federal_excise_tax -- only AZ and a couple others break this out",
            "channel -- some states are online-only, some don't split"
        ]
    }

    output_path = RESEARCH_DIR / "unified_schema_recommendation.json"
    with open(output_path, "w") as f:
        json.dump(schema, f, indent=2)
    print(f"Saved unified schema to {output_path}")
    return output_path

def create_ggr_definitions_md(all_states):
    """Create the GGR definitions markdown document."""
    lines = ["# GGR Definitions by State\n"]
    lines.append("How each state defines and calculates Gross Gaming Revenue (or equivalent) for sports betting.\n")
    lines.append("---\n")

    for code, data in sorted(all_states.items()):
        state_name = safe_get(data, "state_name", code)
        lines.append(f"\n## {state_name} ({code})\n")
        lines.append(f"- **Term used**: {safe_get(data, 'ggr_definition')}")
        lines.append(f"- **Calculation**: {safe_get(data, 'ggr_formula')}")

        promo = safe_get(data, 'promo_deduction')
        lines.append(f"- **Promo deduction**: {promo}")

        lines.append(f"- **Taxable base**: {safe_get(data, 'tax_basis')}")
        lines.append(f"- **Online tax rate**: {safe_get(data, 'online_tax_rate')}")
        lines.append(f"- **Retail tax rate**: {safe_get(data, 'retail_tax_rate')}")

        notes = safe_get(data, "notes", "")
        if notes and notes != "UNKNOWN":
            lines.append(f"- **Notes**: {notes}")
        lines.append("")

    output_path = RESEARCH_DIR / "ggr_definitions_by_state.md"
    with open(output_path, "w") as f:
        f.write("\n".join(lines))
    print(f"Saved GGR definitions to {output_path}")
    return output_path

def update_research_log(all_states):
    """Update the research log summary."""
    total = len(all_states)
    with_samples = sum(1 for d in all_states.values()
                       if d.get("sample_files_downloaded") and len(d.get("sample_files_downloaded", [])) > 0)
    with_operators = sum(1 for d in all_states.values() if d.get("operator_breakdown"))
    with_sports = sum(1 for d in all_states.values() if d.get("sport_breakdown"))
    needs_review = sum(1 for d in all_states.values()
                       if d.get("scraping_difficulty", "").upper() in ["HARD", "NOT FEASIBLE", "UNKNOWN"])

    # Count total files
    samples_dir = RESEARCH_DIR / "samples"
    total_files = sum(len(files) for _, _, files in os.walk(samples_dir)) if samples_dir.exists() else 0

    summary = f"""# Research Log — US State Sports Betting Revenue Data

## Summary
- States researched: {total}/33
- States with sample data downloaded: {with_samples}
- States with operator breakdown confirmed: {with_operators}
- States with sport breakdown confirmed: {with_sports}
- States that need manual review: {needs_review}
- Total files collected: {total_files}
- Research completed: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}

---

"""

    # Read existing log entries (skip old header)
    log_path = RESEARCH_DIR / "research_log.md"
    existing_entries = ""
    if log_path.exists():
        with open(log_path) as f:
            content = f.read()
            # Find first "### " entry
            idx = content.find("\n### ")
            if idx >= 0:
                existing_entries = content[idx:]

    with open(log_path, "w") as f:
        f.write(summary + existing_entries)

    print(f"Updated research log")

def main():
    print("=" * 60)
    print("COMPILING ALL RESEARCH RESULTS")
    print("=" * 60)

    print("\n1. Loading tier results...")
    all_states = load_all_results()
    print(f"   Total states loaded: {len(all_states)}")

    print("\n2. Loading tax/GGR research...")
    tax_data = load_tax_research()
    print(f"   Tax data entries: {len(tax_data)}")

    print("\n3. Merging tax data...")
    all_states = merge_tax_data(all_states, tax_data)
    print(f"   Total states after merge: {len(all_states)}")

    print("\n4. Creating master spreadsheet...")
    create_master_spreadsheet(all_states)

    print("\n5. Creating operator mapping JSON...")
    create_operator_mapping_json(all_states)

    print("\n6. Creating unified schema JSON...")
    create_unified_schema_json()

    print("\n7. Creating GGR definitions document...")
    create_ggr_definitions_md(all_states)

    print("\n8. Updating research log...")
    update_research_log(all_states)

    print("\n" + "=" * 60)
    print("COMPILATION COMPLETE")
    print("=" * 60)

    # List output files
    print("\nOutput files:")
    for f in sorted(RESEARCH_DIR.glob("*")):
        if f.is_file():
            size = f.stat().st_size
            print(f"  {f.name} ({size:,} bytes)")

if __name__ == "__main__":
    main()
