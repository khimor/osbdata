#!/usr/bin/env python3
"""
Final compilation: Merge all tier results + tax research into clean master outputs.
Filters out metadata keys, enriches operator mappings, and generates all deliverables.
"""
import json
import os
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESEARCH_DIR = Path("/Users/nosherzapoo/Desktop/claude/osb-trackerv0/research")

# Valid US state/territory codes
VALID_CODES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","DC","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ","NM",
    "NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT","VA","WA",
    "WV","WI","WY"
}

TIER_MAP = {
    "NY": 1, "IL": 1, "PA": 1, "NJ": 1, "OH": 1, "MI": 1,
    "AZ": 2, "IN": 2, "MA": 2, "MD": 2, "VA": 2, "CO": 2,
    "CT": 3, "KY": 3, "TN": 3, "LA": 3, "NC": 3, "KS": 3, "IA": 3,
    "WV": 4, "ME": 4, "NH": 4, "RI": 4, "WY": 4, "MO": 4,
    "MT": 5, "OR": 5, "SD": 5, "DE": 5, "DC": 5, "AR": 5, "VT": 5, "MS": 5, "NV": 5
}

STATE_NAMES = {
    "NY": "New York", "IL": "Illinois", "PA": "Pennsylvania", "NJ": "New Jersey",
    "OH": "Ohio", "MI": "Michigan", "AZ": "Arizona", "IN": "Indiana",
    "MA": "Massachusetts", "MD": "Maryland", "VA": "Virginia", "CO": "Colorado",
    "CT": "Connecticut", "KY": "Kentucky", "TN": "Tennessee", "LA": "Louisiana",
    "NC": "North Carolina", "KS": "Kansas", "IA": "Iowa", "WV": "West Virginia",
    "ME": "Maine", "NH": "New Hampshire", "RI": "Rhode Island", "WY": "Wyoming",
    "MO": "Missouri", "MT": "Montana", "OR": "Oregon", "SD": "South Dakota",
    "DE": "Delaware", "DC": "District of Columbia", "AR": "Arkansas", "VT": "Vermont",
    "MS": "Mississippi", "NV": "Nevada"
}

# ========================================================================
# OPERATOR STANDARDIZATION DATABASE
# ========================================================================
OPERATOR_DB = {
    "FanDuel": {
        "parent_company": "Flutter Entertainment",
        "notes": "Largest US sportsbook by market share (~40%+). Flutter Entertainment plc (LSE: FLTR)."
    },
    "DraftKings": {
        "parent_company": "DraftKings Inc.",
        "notes": "Second-largest US sportsbook (~25%+ market share). NASDAQ: DKNG."
    },
    "BetMGM": {
        "parent_company": "Entain/MGM Resorts (50/50 JV)",
        "notes": "Joint venture between Entain plc and MGM Resorts International."
    },
    "Caesars Sportsbook": {
        "parent_company": "Caesars Entertainment",
        "notes": "Acquired William Hill in 2021. NASDAQ: CZR."
    },
    "ESPN Bet": {
        "parent_company": "PENN Entertainment",
        "notes": "Formerly Barstool Sportsbook (2020-2023), then rebranded ESPN Bet (Nov 2023). Also absorbed WynnBET license in some states."
    },
    "BetRivers": {
        "parent_company": "Rush Street Interactive",
        "notes": "Also operates as SugarHouse in some markets. NYSE: RSI."
    },
    "Fanatics Sportsbook": {
        "parent_company": "Fanatics Betting & Gaming",
        "notes": "Acquired PointsBet US operations in 2023. Rebranded from PointsBet to Fanatics."
    },
    "bet365": {
        "parent_company": "bet365 Group Ltd",
        "notes": "UK-based, largest online bookmaker globally. Limited US presence expanding."
    },
    "Hard Rock Bet": {
        "parent_company": "Seminole Tribe of Florida / Hard Rock Digital",
        "notes": "Primarily Florida-based with expanding US footprint."
    },
    "Bally Bet": {
        "parent_company": "Bally's Corporation",
        "notes": "Smaller market share. NYSE: BALY."
    },
    "Circa Sports": {
        "parent_company": "Circa Resort & Casino",
        "notes": "Nevada-based, expanding to select markets (IL, CO, IA)."
    },
    "theScore Bet": {
        "parent_company": "PENN Entertainment",
        "notes": "Acquired by PENN in 2021. Now being phased into ESPN Bet brand."
    },
    "Resorts Digital": {
        "parent_company": "Resorts Casino Hotel / DGC (PokerStars)",
        "notes": "Limited to NJ/PA. Affiliated with PokerStars brand."
    },
    "Golden Nugget": {
        "parent_company": "Tilman Fertitta / DraftKings",
        "notes": "Golden Nugget Online Gaming acquired by DraftKings in 2022."
    },
    "Parx": {
        "parent_company": "Greenwood Racing Inc.",
        "notes": "PA-based casino operator. Online branded as betPARX."
    },
    "Betly": {
        "parent_company": "Delaware North",
        "notes": "Operates in AR, WV. Small regional operator."
    },
    "BetSaracen": {
        "parent_company": "Saracen Casino Resort",
        "notes": "Arkansas-only operator."
    },
    "Oaklawn": {
        "parent_company": "Oaklawn Racing Casino Resort",
        "notes": "Arkansas-only, historic racetrack & casino."
    },
    "GambetDC": {
        "parent_company": "Intralot / DC Lottery",
        "notes": "Replaced by FanDuel in 2024."
    },
    "Sports Bet Montana": {
        "parent_company": "Intralot / Montana Lottery",
        "notes": "Sole operator monopoly in Montana."
    },
    "Sportsbook RI": {
        "parent_company": "IGT / Bally's Corporation / RI Lottery",
        "notes": "State-controlled monopoly through Rhode Island Lottery."
    },
    "Scoreboard (OR)": {
        "parent_company": "DraftKings / Oregon Lottery",
        "notes": "Oregon Lottery's sports betting platform operated by DraftKings."
    }
}

# Alias -> Standard Name mapping
ALIAS_MAP = {
    # FanDuel aliases
    "fanduel": "FanDuel", "fanduel sportsbook": "FanDuel", "fanduel group": "FanDuel",
    "flutter/fanduel": "FanDuel", "betfair interactive us, llc d/b/a fanduel sportsbook": "FanDuel",
    "betfair interactive us llc": "FanDuel",
    # DraftKings aliases
    "draftkings": "DraftKings", "draftkings sportsbook": "DraftKings", "draftkings sport book": "DraftKings",
    "crown gaming / draftkings": "DraftKings", "crown gaming/draftkings": "DraftKings",
    "dk crown holdings, inc. d/b/a draftkings": "DraftKings", "dk": "DraftKings",
    # BetMGM aliases
    "betmgm": "BetMGM", "bet mgm": "BetMGM", "roar digital": "BetMGM",
    "betmgm, llc d/b/a betmgm": "BetMGM", "borgata": "BetMGM",
    # Caesars aliases
    "caesars": "Caesars Sportsbook", "caesars sportsbook": "Caesars Sportsbook",
    "caesars sport book": "Caesars Sportsbook", "william hill": "Caesars Sportsbook",
    "caesars interactive": "Caesars Sportsbook",
    "american wagering, inc. d/b/a caesars sportsbook": "Caesars Sportsbook",
    # ESPN Bet / PENN aliases
    "espn bet": "ESPN Bet", "espn bet (penn entertainment)": "ESPN Bet",
    "penn entertainment / espn bet": "ESPN Bet", "penn entertainment": "ESPN Bet",
    "wynn interactive": "ESPN Bet", "wynnbet": "ESPN Bet", "wynn bet": "ESPN Bet",
    "barstool sportsbook": "ESPN Bet", "barstool": "ESPN Bet",
    "barstool/espn bet": "ESPN Bet", "espnbet": "ESPN Bet",
    # BetRivers / RSI aliases
    "betrivers": "BetRivers", "rush street interactive": "BetRivers",
    "rivers": "BetRivers", "sugarhouse": "BetRivers",
    "rsi la, llc d/b/a betrivers sportsbook": "BetRivers",
    # Fanatics aliases
    "fanatics": "Fanatics Sportsbook", "fanatics sportsbook": "Fanatics Sportsbook",
    "fanatics betting": "Fanatics Sportsbook", "pointsbet": "Fanatics Sportsbook",
    # bet365
    "bet365": "bet365",
    # Hard Rock
    "hard rock": "Hard Rock Bet", "hard rock bet": "Hard Rock Bet",
    "hard rock digital": "Hard Rock Bet",
    # Bally
    "bally bet": "Bally Bet", "ballybet": "Bally Bet", "bally's": "Bally Bet",
    "bally bet (bally's interactive, llc)": "Bally Bet",
    # Circa
    "circa sports": "Circa Sports", "circa": "Circa Sports",
    # theScore
    "thescore bet": "theScore Bet", "thescore": "theScore Bet",
    # Other
    "golden nugget": "Golden Nugget", "golden nugget online gaming": "Golden Nugget",
    "resorts digital": "Resorts Digital", "parx casino": "Parx", "parx": "Parx",
    "betparx": "Parx",
    "betly": "Betly",
}

def sg(d, key, default="UNKNOWN"):
    """Safe get."""
    val = d.get(key, default)
    if val is None or val == "" or val == []:
        return default
    return val

def load_all_results():
    """Load and merge all tier results, filtering out non-state keys."""
    all_states = {}
    for tier_file in ["tier1_results.json", "tier2_results.json", "tier3_results.json",
                      "tier4_results.json", "tier5_results.json"]:
        path = RESEARCH_DIR / tier_file
        if path.exists():
            with open(path) as f:
                data = json.load(f)
                for key, val in data.items():
                    # Only keep valid state codes
                    if key.upper() in VALID_CODES and isinstance(val, dict):
                        all_states[key.upper()] = val
            print(f"  Loaded {tier_file}")
    return all_states

def load_tax_research():
    path = RESEARCH_DIR / "tax_and_ggr_research.json"
    if path.exists():
        with open(path) as f:
            data = json.load(f)
        # Filter to valid state codes
        return {k: v for k, v in data.items() if k.upper() in VALID_CODES and isinstance(v, dict)}
    return {}

def merge_tax_data(all_states, tax_data):
    for code, tax_info in tax_data.items():
        code = code.upper()
        if code in all_states:
            state = all_states[code]
            for field in ["online_tax_rate", "retail_tax_rate", "tax_basis", "launch_date"]:
                if not state.get(field) or state[field] in ["UNKNOWN", "", None]:
                    state[field] = tax_info.get(field, state.get(field, "UNKNOWN"))
            if not state.get("ggr_definition") or state["ggr_definition"] in ["UNKNOWN", "", None]:
                state["ggr_definition"] = tax_info.get("ggr_term_used", "UNKNOWN")
            if not state.get("ggr_formula") or state["ggr_formula"] in ["UNKNOWN", "", None]:
                state["ggr_formula"] = tax_info.get("ggr_formula", "UNKNOWN")
            if not state.get("promo_deduction") or state["promo_deduction"] in ["UNKNOWN", "", None]:
                pda = tax_info.get("promo_deduction_allowed", "")
                pdd = tax_info.get("promo_deduction_details", "")
                state["promo_deduction"] = pdd if pdd else str(pda)
            if tax_info.get("special_notes"):
                existing = state.get("notes", "") or ""
                if tax_info["special_notes"] not in existing:
                    state["notes"] = f"{existing} | {tax_info['special_notes']}" if existing else tax_info["special_notes"]
        else:
            all_states[code] = {
                "state_name": STATE_NAMES.get(code, code),
                "regulatory_body": "UNKNOWN",
                "url": "UNKNOWN",
                "url_status": "not_checked",
                "launch_date": tax_info.get("launch_date", "UNKNOWN"),
                "reporting_frequency": "UNKNOWN",
                "file_formats": [],
                "has_archive": False, "archive_depth": "UNKNOWN",
                "requires_js": False,
                "operator_breakdown": False, "operators": [],
                "sport_breakdown": False, "sport_categories": [],
                "retail_online_split": False,
                "data_fields": [],
                "ggr_definition": tax_info.get("ggr_term_used", "UNKNOWN"),
                "ggr_formula": tax_info.get("ggr_formula", "UNKNOWN"),
                "promo_deduction": tax_info.get("promo_deduction_details", "UNKNOWN"),
                "online_tax_rate": tax_info.get("online_tax_rate", "UNKNOWN"),
                "retail_tax_rate": tax_info.get("retail_tax_rate", "UNKNOWN"),
                "tax_basis": tax_info.get("tax_basis", "UNKNOWN"),
                "scraping_difficulty": "UNKNOWN",
                "scraping_approach": "UNKNOWN",
                "scraping_time_estimate": "UNKNOWN",
                "sample_files_downloaded": [],
                "notes": tax_info.get("special_notes", ""),
                "log": "Data from supplementary tax research only"
            }
    return all_states

def normalize_operator(raw_name):
    """Normalize an operator name to its standard form."""
    if not raw_name or not isinstance(raw_name, str):
        return raw_name
    # Try exact match first
    clean = raw_name.strip()
    lower = clean.lower()
    if lower in ALIAS_MAP:
        return ALIAS_MAP[lower]
    # Try partial matches
    for alias, standard in ALIAS_MAP.items():
        if alias in lower:
            return standard
    return clean

def build_operator_mapping(all_states):
    """Build comprehensive operator mapping from all state data."""
    # Track: standard_name -> {aliases, active_states}
    op_data = defaultdict(lambda: {"aliases": set(), "active_states": set()})
    state_mappings = {}

    for code, data in sorted(all_states.items()):
        operators = data.get("operators", [])
        if not isinstance(operators, list):
            continue
        for op in operators:
            if not isinstance(op, str) or not op.strip():
                continue
            raw = op.strip()
            std = normalize_operator(raw)

            if std in OPERATOR_DB:
                op_data[std]["aliases"].add(raw)
                op_data[std]["active_states"].add(code)
                if raw.lower() != std.lower():
                    if code not in state_mappings:
                        state_mappings[code] = {}
                    state_mappings[code][raw] = std
            else:
                # Unknown operator - add as-is
                op_data[std]["aliases"].add(raw)
                op_data[std]["active_states"].add(code)

    output = {"standardized_names": {}, "state_specific_mappings": state_mappings}

    # Known operators first
    for name, info in OPERATOR_DB.items():
        if name in op_data:
            output["standardized_names"][name] = {
                "aliases": sorted(op_data[name]["aliases"]),
                "parent_company": info["parent_company"],
                "active_states": sorted(op_data[name]["active_states"]),
                "notes": info["notes"]
            }
        else:
            output["standardized_names"][name] = {
                "aliases": [name],
                "parent_company": info["parent_company"],
                "active_states": [],
                "notes": info["notes"]
            }

    # Unknown operators (not in OPERATOR_DB)
    for name, data in sorted(op_data.items()):
        if name not in output["standardized_names"]:
            output["standardized_names"][name] = {
                "aliases": sorted(data["aliases"]),
                "parent_company": "Unknown",
                "active_states": sorted(data["active_states"]),
                "notes": ""
            }

    path = RESEARCH_DIR / "operator_name_mapping.json"
    with open(path, "w") as f:
        json.dump(output, f, indent=2)
    print(f"  Saved operator mapping ({len(output['standardized_names'])} operators)")
    return output

def build_spreadsheet(all_states):
    """Build the 7-sheet master Excel spreadsheet."""
    wb = Workbook()

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign
            cell.border = border

    def auto_width(ws, mn=10, mx=45):
        for col in ws.columns:
            ml = 0
            cl = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        ml = max(ml, min(len(str(cell.value)), mx))
                except:
                    pass
            ws.column_dimensions[cl].width = max(ml + 2, mn)

    sorted_states = sorted(all_states.items())

    # ===== Sheet 1: State Overview =====
    ws = wb.active
    ws.title = "State Overview"
    h = ["State Code", "State Name", "Regulatory Body", "URL", "URL Status",
         "Launch Date", "Frequency", "Format", "Operator Breakdown (Y/N)",
         "Sport Breakdown (Y/N)", "Retail/Online Split (Y/N)", "Online Tax Rate",
         "Retail Tax Rate", "Tax Basis", "Scraping Difficulty",
         "Est. Scraper Time", "Archive Depth", "Sample Saved (Y/N)", "Tier", "Notes"]
    ws.append(h)
    style_header(ws, len(h))

    for code, d in sorted_states:
        fmts = sg(d, "file_formats", [])
        if isinstance(fmts, list):
            fmts = ", ".join(str(f) for f in fmts)
        samples = sg(d, "sample_files_downloaded", [])
        has_s = "Y" if isinstance(samples, list) and len(samples) > 0 else "N"

        ws.append([
            code, sg(d, "state_name", STATE_NAMES.get(code, code)),
            sg(d, "regulatory_body"), sg(d, "url"), sg(d, "url_status"),
            sg(d, "launch_date"), sg(d, "reporting_frequency"),
            str(fmts)[:100],
            "Y" if d.get("operator_breakdown") else "N",
            "Y" if d.get("sport_breakdown") else "N",
            "Y" if d.get("retail_online_split") else "N",
            sg(d, "online_tax_rate"), sg(d, "retail_tax_rate"),
            sg(d, "tax_basis"), sg(d, "scraping_difficulty"),
            sg(d, "scraping_time_estimate"), sg(d, "archive_depth"),
            has_s, TIER_MAP.get(code, "?"),
            str(sg(d, "notes", ""))[:200]
        ])
    auto_width(ws)

    # ===== Sheet 2: Data Fields by State =====
    ws2 = wb.create_sheet("Data Fields by State")
    standard_fields = [
        "Handle (Total Wagers)", "Online Handle", "Retail Handle",
        "Gross Revenue (before promos)", "Promo Credits / Free Bets",
        "Net Revenue (after promos) / AGR", "Payouts / Winnings Paid",
        "Hold %", "Tax Paid", "Tax Rate", "Operator Name",
        "Sport Category", "Bet Type (parlay/straight/etc)",
        "New Accounts / Registrations", "Number of Bets / Transactions",
        "Federal Excise Tax"
    ]

    h2 = ["State Code"] + standard_fields
    ws2.append(h2)
    style_header(ws2, len(h2))

    for code, d in sorted_states:
        # Build a set of what this state reports
        state_fields_lower = set()
        fields = d.get("data_fields", [])
        if isinstance(fields, list):
            for f in fields:
                if isinstance(f, dict):
                    state_fields_lower.add(f.get("name", "").lower())
                elif isinstance(f, str):
                    state_fields_lower.add(f.lower())

        row = [code]
        for sf in standard_fields:
            sf_lower = sf.lower()
            keywords = sf_lower.replace("(", "").replace(")", "").replace("/", " ").split()

            found = False
            for sf_entry in state_fields_lower:
                if any(kw in sf_entry for kw in keywords if len(kw) > 3):
                    found = True
                    break

            # Also check structural flags
            if not found:
                if "operator" in sf_lower and d.get("operator_breakdown"):
                    found = True
                elif "sport" in sf_lower and d.get("sport_breakdown"):
                    found = True
                elif "online" in sf_lower and d.get("retail_online_split"):
                    found = True
                elif "retail" in sf_lower and d.get("retail_online_split"):
                    found = True

            row.append("Y" if found else "N")
        ws2.append(row)
    auto_width(ws2, mn=8, mx=30)

    # ===== Sheet 3: Operator Name Mapping =====
    ws3 = wb.create_sheet("Operator Name Mapping")
    h3 = ["State", "Name As Reported", "Standardized Name", "Parent Company"]
    ws3.append(h3)
    style_header(ws3, len(h3))

    for code, d in sorted_states:
        ops = d.get("operators", [])
        if not isinstance(ops, list):
            continue
        for op in ops:
            if not isinstance(op, str) or not op.strip():
                continue
            raw = op.strip()
            std = normalize_operator(raw)
            parent = OPERATOR_DB.get(std, {}).get("parent_company", "")
            ws3.append([code, raw, std, parent])
    auto_width(ws3)

    # ===== Sheet 4: GGR Definitions =====
    ws4 = wb.create_sheet("GGR Definitions")
    h4 = ["State", "GGR Term Used", "Formula Description",
          "Promo Deduction Allowed", "Promo Deduction Limit/Details",
          "Taxable Base Definition", "Notes"]
    ws4.append(h4)
    style_header(ws4, len(h4))

    for code, d in sorted_states:
        promo = sg(d, "promo_deduction", "UNKNOWN")
        promo_yn = "UNKNOWN"
        if isinstance(promo, bool):
            promo_yn = "Y" if promo else "N"
        elif isinstance(promo, str):
            pl = promo.lower()
            if any(x in pl for x in ["not deduct", "not allowed", "no deduct", "no promo"]):
                promo_yn = "N"
            elif any(x in pl for x in ["fully", "unlimited", "yes", "allowed", "deductible"]):
                promo_yn = "Y"
            elif any(x in pl for x in ["partial", "sliding", "limited", "cap", "first 12"]):
                promo_yn = "Partial"

        ws4.append([
            code, sg(d, "ggr_definition"), sg(d, "ggr_formula"),
            promo_yn, str(promo)[:150], sg(d, "tax_basis"),
            str(sg(d, "notes", ""))[:200]
        ])
    auto_width(ws4)

    # ===== Sheet 5: Sport Categories =====
    ws5 = wb.create_sheet("Sport Categories")
    h5 = ["State", "Sport Category", "Handle Reported", "Revenue Reported"]
    ws5.append(h5)
    style_header(ws5, len(h5))

    for code, d in sorted_states:
        if d.get("sport_breakdown"):
            cats = d.get("sport_categories", [])
            if isinstance(cats, list) and len(cats) > 0:
                for cat in cats:
                    if isinstance(cat, str):
                        ws5.append([code, cat, "Y", "Y"])
                    elif isinstance(cat, dict):
                        ws5.append([code, cat.get("name", str(cat)),
                                    "Y" if cat.get("handle_reported", True) else "N",
                                    "Y" if cat.get("revenue_reported", True) else "N"])
            else:
                ws5.append([code, "(sport breakdown confirmed but categories not documented)", "UNKNOWN", "UNKNOWN"])
    auto_width(ws5)

    # ===== Sheet 6: Scraping Plan =====
    ws6 = wb.create_sheet("Scraping Plan")
    h6 = ["State", "Recommended Tool", "URL Pattern", "Auth/Headers",
          "Anti-Scraping Notes", "Est. Build Time", "Priority"]
    ws6.append(h6)
    style_header(ws6, len(h6))

    priority = 1
    for tier_codes in [
        ["NY", "IL", "PA", "NJ", "OH", "MI"],
        ["AZ", "IN", "MA", "MD", "VA", "CO"],
        ["CT", "KY", "TN", "LA", "NC", "KS", "IA"],
        ["WV", "ME", "NH", "RI", "WY", "MO"],
        ["MT", "OR", "SD", "DE", "DC", "AR", "VT", "MS", "NV"]
    ]:
        for code in tier_codes:
            d = all_states.get(code, {})
            auth = "Playwright needed" if d.get("requires_js") else "Standard HTTP"
            ws6.append([
                code, sg(d, "scraping_approach"), sg(d, "url"),
                auth, str(sg(d, "notes", ""))[:100],
                sg(d, "scraping_time_estimate"), priority
            ])
            priority += 1
    auto_width(ws6)

    # ===== Sheet 7: Sample Data Preview =====
    ws7 = wb.create_sheet("Sample Data Preview")
    h7 = ["State", "Source File", "Format", "Key Fields Found"]
    ws7.append(h7)
    style_header(ws7, len(h7))

    for code, d in sorted_states:
        samples = d.get("sample_files_downloaded", [])
        fields = d.get("data_fields", [])
        field_str = ""
        if isinstance(fields, list):
            field_names = []
            for f in fields[:8]:
                if isinstance(f, dict):
                    field_names.append(f.get("name", ""))
                elif isinstance(f, str):
                    field_names.append(f)
            field_str = ", ".join(fn for fn in field_names if fn)

        if isinstance(samples, list) and len(samples) > 0:
            for sf in samples[:2]:
                if isinstance(sf, str):
                    ext = sf.rsplit(".", 1)[-1] if "." in sf else "unknown"
                    ws7.append([code, os.path.basename(sf), ext.upper(), field_str[:200]])
                elif isinstance(sf, dict):
                    ws7.append([code, sf.get("filename", "unknown"),
                                sf.get("format", "unknown"), field_str[:200]])
        else:
            fmts = d.get("file_formats", [])
            fmt_str = ", ".join(str(f) for f in fmts) if isinstance(fmts, list) else str(fmts)
            ws7.append([code, "(no sample downloaded)", fmt_str, field_str[:200]])
    auto_width(ws7)

    path = RESEARCH_DIR / "state_data_research.xlsx"
    wb.save(path)
    print(f"  Saved master spreadsheet ({path.stat().st_size:,} bytes)")

def build_ggr_doc(all_states, tax_data):
    """Build comprehensive GGR definitions markdown."""
    lines = ["# GGR Definitions by State",
             "",
             "How each state defines and calculates Gross Gaming Revenue (or equivalent) for sports betting.",
             "",
             "---",
             ""]

    for code in sorted(all_states.keys()):
        d = all_states[code]
        td = tax_data.get(code, {})
        name = sg(d, "state_name", STATE_NAMES.get(code, code))

        lines.append(f"## {name} ({code})")
        lines.append("")

        # GGR term
        term = sg(d, "ggr_definition")
        if term == "UNKNOWN" and td:
            term = td.get("ggr_term_used", "UNKNOWN")
        lines.append(f"- **Term used**: {term}")

        # Formula
        formula = sg(d, "ggr_formula")
        if formula == "UNKNOWN" and td:
            formula = td.get("ggr_formula", "UNKNOWN")
        lines.append(f"- **Calculation**: {formula}")

        # Promo deduction
        promo = sg(d, "promo_deduction")
        if promo == "UNKNOWN" and td:
            promo = td.get("promo_deduction_details", "UNKNOWN")
        lines.append(f"- **Promo deduction**: {promo}")

        # Tax rates
        online_rate = sg(d, "online_tax_rate")
        if online_rate == "UNKNOWN" and td:
            online_rate = td.get("online_tax_rate", "UNKNOWN")
        retail_rate = sg(d, "retail_tax_rate")
        if retail_rate == "UNKNOWN" and td:
            retail_rate = td.get("retail_tax_rate", "UNKNOWN")

        lines.append(f"- **Online tax rate**: {online_rate}")
        lines.append(f"- **Retail tax rate**: {retail_rate}")

        # Tax basis
        basis = sg(d, "tax_basis")
        if basis == "UNKNOWN" and td:
            basis = td.get("tax_basis", "UNKNOWN")
        lines.append(f"- **Taxable base**: {basis}")

        # Special notes from tax research
        notes = td.get("special_notes", "")
        if notes:
            lines.append(f"- **Why it matters**: {notes}")

        lines.append("")

    path = RESEARCH_DIR / "ggr_definitions_by_state.md"
    with open(path, "w") as f:
        f.write("\n".join(lines))
    print(f"  Saved GGR definitions ({len(all_states)} states)")

def build_unified_schema():
    """Build the unified schema recommendation JSON."""
    schema = {
        "description": "Recommended database schema that accommodates all US state sports betting revenue data",
        "tables": {
            "states": {
                "description": "Reference table for state-level configuration",
                "columns": {
                    "state_code": "VARCHAR(2) PRIMARY KEY",
                    "state_name": "VARCHAR(50) NOT NULL",
                    "regulatory_body": "VARCHAR(100)",
                    "source_url": "TEXT",
                    "launch_date": "DATE",
                    "reporting_frequency": "VARCHAR(20) -- 'weekly', 'monthly', 'quarterly'",
                    "online_tax_rate": "DECIMAL(5,4)",
                    "retail_tax_rate": "DECIMAL(5,4)",
                    "tax_basis": "VARCHAR(20) -- 'ggr', 'handle', 'agr', 'revenue_share'",
                    "promo_deduction_allowed": "BOOLEAN",
                    "promo_deduction_cap": "TEXT -- description of any cap/limit",
                    "ggr_definition": "TEXT -- how this state defines GGR",
                    "data_format": "VARCHAR(50) -- 'xlsx', 'csv', 'pdf', 'html'",
                    "has_operator_breakdown": "BOOLEAN",
                    "has_sport_breakdown": "BOOLEAN",
                    "has_channel_split": "BOOLEAN"
                }
            },
            "monthly_data": {
                "description": "Main data table — one row per state/period/operator/channel/sport combination",
                "columns": {
                    "id": "BIGSERIAL PRIMARY KEY",
                    "state_code": "VARCHAR(2) NOT NULL REFERENCES states(state_code)",
                    "period_start": "DATE NOT NULL",
                    "period_end": "DATE NOT NULL",
                    "period_type": "VARCHAR(10) NOT NULL -- 'weekly', 'monthly'",
                    "operator_raw": "VARCHAR(200) -- name exactly as reported by state",
                    "operator_standard": "VARCHAR(100) -- normalized name from mapping table",
                    "channel": "VARCHAR(20) NOT NULL DEFAULT 'combined' -- 'online', 'retail', 'kiosk', 'combined'",
                    "sport_category": "VARCHAR(50) -- NULL if state doesn't break down by sport",
                    "handle": "BIGINT -- total wagers in cents",
                    "gross_revenue": "BIGINT -- handle minus payouts, before promo deductions, in cents",
                    "promo_credits": "BIGINT -- free bets / promotional credits in cents, NULL if not reported",
                    "net_revenue": "BIGINT -- after promo deductions (taxable base in most states), in cents",
                    "payouts": "BIGINT -- winnings paid to bettors, NULL if not reported, in cents",
                    "hold_pct": "DECIMAL(6,4) -- gross_revenue / handle, NULL if not calculable",
                    "tax_paid": "BIGINT -- actual tax remitted, in cents",
                    "num_bets": "INTEGER -- number of wagers, NULL if not reported",
                    "federal_excise_tax": "BIGINT -- NULL if not reported separately, in cents",
                    "created_at": "TIMESTAMP NOT NULL DEFAULT NOW()",
                    "source_file": "TEXT -- filename/URL of source report"
                },
                "indexes": [
                    "CREATE UNIQUE INDEX idx_monthly_unique ON monthly_data(state_code, period_end, COALESCE(operator_standard,''), channel, COALESCE(sport_category,''))",
                    "CREATE INDEX idx_monthly_state_period ON monthly_data(state_code, period_end)",
                    "CREATE INDEX idx_monthly_operator ON monthly_data(operator_standard, period_end)"
                ],
                "notes": "Store currency as cents (BIGINT) to avoid floating point issues. Convert to dollars in application layer. Use COALESCE in unique index for nullable sport_category."
            },
            "operators": {
                "description": "Operator name normalization reference table",
                "columns": {
                    "id": "SERIAL PRIMARY KEY",
                    "standard_name": "VARCHAR(100) UNIQUE NOT NULL",
                    "parent_company": "VARCHAR(100)",
                    "notes": "TEXT"
                }
            },
            "operator_aliases": {
                "description": "Maps raw operator names to standardized names",
                "columns": {
                    "id": "SERIAL PRIMARY KEY",
                    "state_code": "VARCHAR(2) NOT NULL",
                    "raw_name": "VARCHAR(200) NOT NULL",
                    "standard_name": "VARCHAR(100) NOT NULL REFERENCES operators(standard_name)",
                    "UNIQUE": "(state_code, raw_name)"
                }
            },
            "scrape_log": {
                "description": "Audit log for data collection runs",
                "columns": {
                    "id": "SERIAL PRIMARY KEY",
                    "state_code": "VARCHAR(2) NOT NULL",
                    "scrape_timestamp": "TIMESTAMP NOT NULL DEFAULT NOW()",
                    "status": "VARCHAR(20) NOT NULL -- 'success', 'partial', 'failed'",
                    "rows_inserted": "INTEGER DEFAULT 0",
                    "rows_updated": "INTEGER DEFAULT 0",
                    "date_range_start": "DATE",
                    "date_range_end": "DATE",
                    "error_message": "TEXT",
                    "source_file": "TEXT",
                    "duration_seconds": "INTEGER"
                }
            }
        },
        "derived_views": {
            "v_hold_pct": "gross_revenue::NUMERIC / NULLIF(handle, 0) AS hold_pct",
            "v_yoy_handle_change": """
                handle - LAG(handle) OVER (
                    PARTITION BY state_code, operator_standard, channel
                    ORDER BY period_end
                ) AS handle_change_vs_prior_period
            """,
            "v_market_share": "handle::NUMERIC / NULLIF(SUM(handle) OVER (PARTITION BY state_code, period_end), 0) AS market_share",
            "v_effective_tax_rate": "tax_paid::NUMERIC / NULLIF(net_revenue, 0) AS effective_tax_rate",
            "v_promo_pct_of_gross": "promo_credits::NUMERIC / NULLIF(gross_revenue, 0) AS promo_pct"
        },
        "columns_that_vary_by_state": [
            "promo_credits -- some states don't report this at all (TN, SD, MT)",
            "payouts -- only ~60% of states report raw payouts separately",
            "sport_category -- only IL, NJ, LA, MO, CO, SD, MT, MS, MD have sport breakdown",
            "num_bets -- very few states (none in standard reports)",
            "federal_excise_tax -- only AZ, ME, KY break this out",
            "channel -- some states are online-only (TN, NC, WY, VT), some don't split"
        ],
        "data_quality_notes": [
            "TN taxes handle (not GGR), so gross_revenue field will be NULL for TN",
            "NY promo credits are NOT deductible, so net_revenue = gross_revenue for NY",
            "RI/NH/MT/OR are lottery monopolies - single operator per state",
            "NV embeds sports in broader gaming report - sports data isolation may be approximate",
            "MO allows 25% promo deduction, causing negative AGR in early months",
            "IL graduated tax tiers (20-40%) + per-wager surcharge starting July 2025"
        ]
    }

    path = RESEARCH_DIR / "unified_schema_recommendation.json"
    with open(path, "w") as f:
        json.dump(schema, f, indent=2)
    print(f"  Saved unified schema recommendation")

def build_research_log(all_states):
    """Build comprehensive research log."""
    # Count stats
    total = len(all_states)
    with_samples = sum(1 for d in all_states.values()
                       if isinstance(d.get("sample_files_downloaded"), list) and len(d["sample_files_downloaded"]) > 0)
    with_operators = sum(1 for d in all_states.values() if d.get("operator_breakdown"))
    with_sports = sum(1 for d in all_states.values() if d.get("sport_breakdown"))

    diff_map = {"HARD": True, "NOT FEASIBLE": True, "UNKNOWN": True}
    needs_review = sum(1 for d in all_states.values()
                       if str(d.get("scraping_difficulty", "")).upper() in diff_map)

    # Count files
    samples_dir = RESEARCH_DIR / "samples"
    total_files = 0
    for root, dirs, files in os.walk(samples_dir):
        total_files += len(files)

    screenshots_dir = RESEARCH_DIR / "screenshots"
    total_screenshots = len(list(screenshots_dir.glob("*.png"))) if screenshots_dir.exists() else 0

    lines = [
        "# Research Log — US State Sports Betting Revenue Data",
        "",
        "## Summary",
        f"- States researched: {total}/33",
        f"- States with sample data downloaded: {with_samples}",
        f"- States with operator breakdown confirmed: {with_operators}",
        f"- States with sport breakdown confirmed: {with_sports}",
        f"- States that need manual review: {needs_review}",
        f"- Total sample files collected: {total_files}",
        f"- Total screenshots captured: {total_screenshots}",
        f"- Research completed: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}",
        "",
        "---",
        ""
    ]

    # Per-state entries
    for code in sorted(all_states.keys()):
        d = all_states[code]
        name = sg(d, "state_name", STATE_NAMES.get(code, code))
        url = sg(d, "url")
        url_status = sg(d, "url_status")
        log_text = sg(d, "log", "")
        difficulty = sg(d, "scraping_difficulty")
        formats = sg(d, "file_formats", [])
        if isinstance(formats, list):
            formats = ", ".join(str(f) for f in formats)

        samples = d.get("sample_files_downloaded", [])
        sample_count = len(samples) if isinstance(samples, list) else 0

        fields = d.get("data_fields", [])
        field_count = len(fields) if isinstance(fields, list) else 0

        lines.append(f"### {code} — {name}")
        lines.append(f"- URL: {url} ({url_status})")
        lines.append(f"- Formats: {formats}")
        lines.append(f"- Sample files downloaded: {sample_count}")
        lines.append(f"- Data fields documented: {field_count}")
        lines.append(f"- Operator breakdown: {'YES' if d.get('operator_breakdown') else 'NO'}")
        lines.append(f"- Sport breakdown: {'YES' if d.get('sport_breakdown') else 'NO'}")
        lines.append(f"- Scraping difficulty: {difficulty}")

        if log_text and log_text != "UNKNOWN":
            lines.append(f"- Notes: {str(log_text)[:300]}")

        lines.append("")

    path = RESEARCH_DIR / "research_log.md"
    with open(path, "w") as f:
        f.write("\n".join(lines))
    print(f"  Updated research log ({total} states)")

def main():
    print("=" * 60)
    print("FINAL COMPILATION — All Research Results")
    print("=" * 60)

    print("\n1. Loading results...")
    all_states = load_all_results()
    print(f"   States loaded: {len(all_states)}")

    print("\n2. Loading tax research...")
    tax_data = load_tax_research()
    print(f"   Tax entries: {len(tax_data)}")

    print("\n3. Merging data...")
    all_states = merge_tax_data(all_states, tax_data)
    # Filter to only states in our target list
    target_states = set(TIER_MAP.keys())
    all_states = {k: v for k, v in all_states.items() if k in target_states}
    print(f"   Final state count: {len(all_states)}")

    print("\n4. Building master spreadsheet...")
    build_spreadsheet(all_states)

    print("\n5. Building operator mapping...")
    build_operator_mapping(all_states)

    print("\n6. Building unified schema...")
    build_unified_schema()

    print("\n7. Building GGR definitions...")
    build_ggr_doc(all_states, tax_data)

    print("\n8. Building research log...")
    build_research_log(all_states)

    print("\n" + "=" * 60)
    print("DONE! Output files:")
    print("=" * 60)
    for f in sorted(RESEARCH_DIR.glob("*.*")):
        if f.is_file():
            print(f"  {f.name:45s} {f.stat().st_size:>10,} bytes")

if __name__ == "__main__":
    main()
